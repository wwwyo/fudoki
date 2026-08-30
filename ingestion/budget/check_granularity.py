"""**調査スクリプト**。本番の Extract ではない。

東京都オープンデータカタログに、予算がどの粒度で出ているかを調べる。
本番の取得は団体ごとに pipeline を作り、届かない団体は最終的に公式サイトの PDF から取る。
これはその前段として「どの団体がどこまで届いているか」を測るためのもの。

判定は列構成で行う（資料名では判定しない。原則3）。判定規則は `granularity_profile.py` が持つ。

⚠️ **母集団は3度も取り違えている。**

1. `data/budget/opendata.json` の**代表1件**だけを見て「事業単位のデータは0件」と判定した
2. organization が `t` + 6桁なら何でも団体として数え、都の部局や都外まで混ぜた（42件中25件が圏外）
3. **全文検索のクエリ語**（歳出 / 当初予算 / 決算書 / 予算データ）で候補を集めた。
   これだと候補が出るのは8団体で、実際に予算資料を持つ17団体の半分も見えない。
   さらに **CSV だけを見ていた**ので、XLSX / XLS でしか出していない9団体が
   丸ごと「予算データ無し」に見えていた（2026-08-30 実測で判明）

だから**団体ごとに `fq=organization:` で全件を列挙する**。クエリ語に依存させない。
「無い」と言えるのは、その団体の全データセットを見たときだけである。

出力はこのスクリプトの隣 `observations/`（ローカル作業ファイル。commit しない）。
**全候補と失敗理由だけを持つ**。団体ごとの最良は観測から導けるので焼き込まない。

⚠️ **ネットワークを叩くのでサンドボックスを外して回す**（AGENTS.md）。

    uv run python -m ingestion.budget.check_granularity            # 結果を表示
    uv run python -m ingestion.budget.check_granularity --write    # observations へ書き出す
"""

from __future__ import annotations

import csv
import io
import json
import pathlib
import sys
import urllib.parse
import zipfile
from collections import Counter
from concurrent.futures import ThreadPoolExecutor

from ingestion.budget.granularity_profile import (
    GRANULARITY_RANK,
    READABLE_FORMATS,
    RELEVANT_TITLE_WORDS,
    classify_granularity,
    detect_direction,
    normalize_column,
    score_header_row,
)
from ingestion.budget.sources import load_catalogs
from ingestion.lib.ckan import datasets_of_organization
from ingestion.lib.http import http_get

ROOT = pathlib.Path(__file__).resolve().parent.parent.parent
# ⚠️ **カタログの宛先をここで宣言しない。** 取得元の正本は sources.toml の [catalog.*] で、
# 団体コードの解決規則（`t` + 6桁）もそこが持つ。写すと、カタログを足したり
# 宛先が変わったりしたときに調査だけが古いカタログを見続ける。
CATALOG = "tokyo"
JURISDICTIONS = ROOT / "ingestion" / "shared" / "jurisdictions.json"
OBSERVATIONS = pathlib.Path(__file__).resolve().parent / "observations"
OUT = OBSERVATIONS / "budget-granularity.json"
# 相手は自治体のサーバ。ホストは団体ごとに分散しているので、この程度なら同一ホストへ集中しない
CONCURRENCY = 3
# 見出しを探す行数。表題・所管課・単位の行が先に来る資料があるので1行目に限らない
HEADER_SCAN_ROWS = 40
# 走査するシート数。年度別に何十枚も持つブックがあるので上限を置く
MAX_SHEETS = 30


def load_jurisdictions() -> dict[str, dict]:
    """母集団。**③会議録のゲート判定は読まない** — 根拠が違ううえ、③が落ちたら①も動かなくなる。"""
    return json.loads(JURISDICTIONS.read_text())["jurisdictions"]


def candidates_of(code: str, packages: list[dict]) -> list[dict]:
    """予算資料の候補。**形式で CSV に絞らない**（XLSX / XLS しか出さない団体が実在する）。"""
    out = []
    for p in packages:
        for r in p.get("resources", []):
            fmt = (r.get("format") or "").upper()
            title = f"{p.get('title', '')} {r.get('name') or ''}"
            if fmt not in (*READABLE_FORMATS, "ZIP"):
                continue
            if not any(k in title for k in RELEVANT_TITLE_WORDS):
                continue
            out.append({
                "code": code,
                "dataset": p.get("title", ""),
                "resource": r.get("name") or "",
                "format": fmt,
                "url": r["url"],
                "direction": detect_direction(title),
                "license_id": p.get("license_id"),
            })
    return out


def _rows_from_csv(body: bytes) -> list[list[str]]:
    for encoding in ("utf-8-sig", "cp932", "utf-8"):
        try:
            text = body.decode(encoding)
        except UnicodeDecodeError:
            continue
        return list(csv.reader(io.StringIO(text)))[:HEADER_SCAN_ROWS]
    raise RuntimeError("復号できる文字コードが無い")


def _rows_from_xlsx(body: bytes) -> list[list[str]]:
    import openpyxl  # noqa: PLC0415  (調査スクリプトなので取得だけのときに import させない)

    book = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    rows: list[list[str]] = []
    for name in book.sheetnames[:MAX_SHEETS]:
        sheet = book[name]
        rows.extend(list(r) for r in sheet.iter_rows(max_row=HEADER_SCAN_ROWS, values_only=True))
    return rows


def _rows_from_xls(body: bytes) -> list[list[str]]:
    import xlrd  # noqa: PLC0415

    book = xlrd.open_workbook(file_contents=body)
    rows: list[list[str]] = []
    for i in range(min(MAX_SHEETS, book.nsheets)):
        sheet = book.sheet_by_index(i)
        rows.extend(sheet.row_values(j) for j in range(min(HEADER_SCAN_ROWS, sheet.nrows)))
    return rows


def _rows_of(fmt: str, body: bytes) -> list[list[str]]:
    if fmt == "CSV":
        return _rows_from_csv(body)
    if fmt in ("XLSX", "XLSM"):
        return _rows_from_xlsx(body)
    if fmt == "XLS":
        # ⚠️ 拡張子と中身が食い違う配信が実在する（`.xls` の名前で XLSX が置いてある）
        try:
            return _rows_from_xls(body)
        except Exception:  # noqa: BLE001
            return _rows_from_xlsx(body)
    raise RuntimeError(f"列を読めない形式: {fmt}")


def _zip_members(body: bytes) -> list[tuple[str, str, bytes]]:
    """ZIP の中の測れるファイル。⚠️ **エントリ名は CP437 で入っている**（Windows 製の書庫）。"""
    out = []
    with zipfile.ZipFile(io.BytesIO(body)) as z:
        for info in z.infolist():
            try:
                name = info.filename.encode("cp437").decode("cp932")
            except (UnicodeEncodeError, UnicodeDecodeError):
                name = info.filename
            fmt = name.rsplit(".", 1)[-1].upper() if "." in name else ""
            if fmt in READABLE_FORMATS:
                out.append((name, fmt, z.read(info)))
    return out


def best_header(rows: list[list[str]]) -> list[str]:
    """見出しらしい行を選ぶ。**1行目とは限らない**（表題・所管課・単位の行が先に来る）。"""
    best: list[str] = []
    score = -1
    for row in rows:
        cells = [c for c in (normalize_column(c) for c in row) if c]
        s = score_header_row(cells)
        if s > score:
            best, score = cells, s
    return best


def observe(candidate: dict) -> dict:
    base = dict(candidate)
    try:
        got = http_get(candidate["url"])
    except Exception as e:  # noqa: BLE001  (取得元の異常は観測として残す。止めない)
        return {**base, "status": None, "granularity": "unchecked", "note": f"{type(e).__name__}: {e}"}
    if got.status != 200:
        return {**base, "status": got.status, "granularity": "unchecked", "note": f"HTTP {got.status}"}

    base |= {"status": got.status, "bytes": len(got.body), "sha256": got.sha256}
    try:
        if candidate["format"] == "ZIP":
            members = _zip_members(got.body)
            if not members:
                return {**base, "granularity": "unchecked", "note": "ZIP に測れるファイルが無い"}
            results = []
            for name, fmt, member in members:
                header = best_header(_rows_of(fmt, member))
                results.append((classify_granularity(header, candidate["code"]), header, name))
            got_best = max(results, key=lambda r: GRANULARITY_RANK.get(r[0].granularity, 0))
            result, header, name = got_best
            base |= {"zip_member": name, "zip_members": len(members)}
        else:
            header = best_header(_rows_of(candidate["format"], got.body))
            result = classify_granularity(header, candidate["code"])
    except Exception as e:  # noqa: BLE001
        return {**base, "granularity": "unchecked", "note": f"読めない: {type(e).__name__}: {e}"}

    return {
        **base,
        "columns": header[:40],
        "granularity": result.granularity,
        "basis": result.basis,
        "note": f"判定根拠の列: {', '.join(result.hits)}" if result.hits else "粒度を示す列が見つからない",
    }


def best_by_jurisdiction(observations: list[dict]) -> dict[str, dict]:
    """団体ごとの最良は観測から導く。ファイルには焼き込まない。

    ⚠️ **歳出だと分かった資料だけを代表にする。** 測っているのは「歳出がどこまで
    届いているか」なので、歳入はもちろん、歳出か歳入か判別できない資料も代表にしない。
    以前は `unknown` を通しており、財政指標の表（「財政力の状況」「財政健全化判断比率」）が
    5団体の代表として歳出の粒度の欄に並んでいた。**測っていないものを結果として出さない。**

    判定できなかったもの（`unchecked`）も同じ理由で代表にしない。
    """
    best: dict[str, dict] = {}
    for o in observations:
        if o["direction"] != "expenditure" or o["granularity"] == "unchecked":
            continue
        prev = best.get(o["code"])
        if prev and GRANULARITY_RANK[prev["granularity"]] >= GRANULARITY_RANK[o["granularity"]]:
            continue
        best[o["code"]] = o
    return best


def main() -> None:
    write = "--write" in sys.argv[1:]
    registry = load_jurisdictions()

    catalog = load_catalogs()[CATALOG]
    candidates: list[dict] = []
    dataset_counts: dict[str, int] = {}
    for code in registry:
        packages = datasets_of_organization(catalog.endpoint, catalog.org_prefix + code)
        dataset_counts[code] = len(packages)
        candidates.extend(candidates_of(code, packages))

    covered = {c["code"] for c in candidates}
    print(f"候補: {len(candidates)} 件 / {len(covered)} 団体（母集団 {len(registry)}）\n")

    with ThreadPoolExecutor(max_workers=CONCURRENCY) as pool:
        observations = list(pool.map(observe, candidates))

    best = best_by_jurisdiction(observations)
    tally = Counter(b["granularity"] for b in best.values())
    for code, b in sorted(best.items()):
        mark = {"project": "✓", "account-item": "◎"}.get(b["granularity"], "△")
        basis = "宣言" if b.get("basis") == "declared" else "推定"
        print(f"  {mark} {code} {registry[code]['name']:8s} {b['granularity']:13s} "
              f"{b['format']:5s} {basis}  {b['dataset'][:30]}")

    # ⚠️ **「歳出として判定できなかった」の内訳を出す。** 候補が1件も無い団体と、
    # 候補はあるが歳出だと分からない（あるいは列から粒度を測れない）団体は別の話で、
    # 前者は「カタログに予算資料が無い」、後者は「資料はあるが歳出の粒度に届かない」。
    with_candidates = {o["code"] for o in observations}
    undecided = with_candidates - set(best)
    print(f"\n{' / '.join(f'{k} {v}' for k, v in sorted(tally.items()))}  "
          f"（歳出の粒度を測れた {len(best)} 団体 / 候補はあるが測れない {len(undecided)} 団体 / "
          f"候補が無い {len(registry) - len(with_candidates)} 団体 / 母集団 {len(registry)}）")
    print(f"全候補: {len(observations)} 件")

    if write:
        OBSERVATIONS.mkdir(parents=True, exist_ok=True)
        OUT.write_text(json.dumps({
            "note": "カタログに予算がどの粒度で出ているかの調査。判定は資料名ではなく列構成で行う（原則3）。"
                    "母集団は団体registry のコードに限り、団体ごとに全データセットを列挙する（原則4）。"
                    "団体ごとの最良はこの観測から導けるので焼き込まない。",
            "generatedBy": "ingestion/budget/check_granularity.py",
            "population": len(registry),
            "datasetCounts": dataset_counts,
            "observations": observations,
        }, ensure_ascii=False, indent=2) + "\n")
        print(f"\n{OUT} へ書き出した")


if __name__ == "__main__":
    main()
